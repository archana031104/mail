import os
import logging
import tempfile
import asyncio
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

logger = logging.getLogger(__name__)

class ExcelProcessor:
    def __init__(self):
        self.screenshot_folder = 'screenshots'
        os.makedirs(self.screenshot_folder, exist_ok=True)
    
    def get_content_range(self, filepath):
        """
        Determine the actual content range in the Excel file
        Returns tuple (last_row, last_col) or None if error
        """
        try:
            # Try with openpyxl first
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            worksheet = workbook.active
            
            # Find the last row and column with data
            last_row = 0
            last_col = 0
            
            for row in worksheet.iter_rows():
                row_num = row[0].row
                for cell in row:
                    if cell.value is not None:
                        last_row = max(last_row, row_num)
                        last_col = max(last_col, cell.column)
            
            workbook.close()
            
            logger.info(f"Content range detected: A1:{get_column_letter(last_col)}{last_row}")
            return last_row, last_col
            
        except Exception as e:
            logger.error(f"Error determining content range with openpyxl: {e}")
            
            # Fallback to pandas
            try:
                df = pd.read_excel(filepath, sheet_name=0)
                last_row = len(df) + 1  # +1 for header
                last_col = len(df.columns)
                
                logger.info(f"Content range detected with pandas: A1:{get_column_letter(last_col)}{last_row}")
                return last_row, last_col
                
            except Exception as e2:
                logger.error(f"Error determining content range with pandas: {e2}")
                return None, None
    
    def create_html_from_excel(self, filepath):
        """
        Convert Excel content to HTML for screenshot
        """
        try:
            # Read Excel file
            df = pd.read_excel(filepath, sheet_name=0)
            
            # Convert to HTML with styling
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        margin: 20px;
                        background-color: white;
                    }}
                    table {{
                        border-collapse: collapse;
                        width: 100%;
                        background-color: white;
                    }}
                    th, td {{
                        border: 1px solid #ddd;
                        padding: 8px;
                        text-align: left;
                        white-space: nowrap;
                    }}
                    th {{
                        background-color: #f2f2f2;
                        font-weight: bold;
                    }}
                    tr:nth-child(even) {{
                        background-color: #f9f9f9;
                    }}
                </style>
            </head>
            <body>
                {df.to_html(index=False, escape=False, table_id='excel-table')}
            </body>
            </html>
            """
            
            return html_content
            
        except Exception as e:
            logger.error(f"Error creating HTML from Excel: {e}")
            return None
    
    async def take_screenshot_async(self, html_content, output_path):
        """
        Take screenshot of HTML content using Playwright
        """
        try:
            async with async_playwright() as p:
                # Launch browser (let Playwright find the installed browser)
                browser = await p.chromium.launch(
                    headless=True
                    # No executable_path - let Playwright use installed browser
                )
                page = await browser.new_page()
                
                # Set viewport for better rendering
                await page.set_viewport_size({"width": 1200, "height": 800})
                
                # Set content
                await page.set_content(html_content)
                
                # Wait for content to load
                await page.wait_for_selector('#excel-table', timeout=10000)
                
                # Get the table element
                table_element = await page.query_selector('#excel-table')
                
                if table_element:
                    # Take screenshot of just the table
                    await table_element.screenshot(path=output_path)
                    logger.info(f"Screenshot saved to: {output_path}")
                    await browser.close()
                    return True
                else:
                    logger.error("Could not find table element for screenshot")
                    await browser.close()
                    return False
                    
        except Exception as e:
            logger.error(f"Error taking screenshot: {e}")
            return False
    
    def process_excel_and_screenshot(self, filepath):
        """
        Main method to process Excel file and take screenshot
        Returns path to screenshot file or None if error
        """
        try:
            # Generate unique filename for screenshot
            base_filename = os.path.splitext(os.path.basename(filepath))[0]
            screenshot_filename = f"{base_filename}_{os.urandom(4).hex()}.png"
            screenshot_path = os.path.join(self.screenshot_folder, screenshot_filename)
            
            # Create HTML from Excel
            html_content = self.create_html_from_excel(filepath)
            
            if not html_content:
                logger.error("Failed to create HTML from Excel")
                return None
            
            # Take screenshot
            success = asyncio.run(self.take_screenshot_async(html_content, screenshot_path))
            
            if success and os.path.exists(screenshot_path):
                return screenshot_path
            else:
                logger.error("Screenshot was not created successfully")
                return None
                
        except Exception as e:
            logger.error(f"Error in process_excel_and_screenshot: {e}")
            return None
